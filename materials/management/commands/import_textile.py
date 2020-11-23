from django.core.management.base import BaseCommand
import openpyxl
import csv
import datetime
from materials.models import TextileManufact, TextileCollection, Textile
from materials.utils import transliterate
import os



class Command(BaseCommand):
    help = 'activate all unactive samples'

    def handle(self, *args, **options):
        print(os.getcwd())
        wb = openpyxl.load_workbook(filename='materials/cat.xlsx')
        sheetnames = wb.sheetnames

        for s in sheetnames:
            sheet = wb[s]
            row_count = sheet.max_row
            print(sheet, row_count)
            i = 2
            y = 2
            z = 2
            max_count = int(row_count)
            while i < max_count:
                supplier = sheet[f'A{i}'].value
                collection = sheet[f'B{i}'].value
                designation = sheet[f'D{i}'].value
                color = sheet[f'F{i}'].value
                model = sheet[f'G{i}'].value
                height = sheet[f'H{i}'].value
                price = sheet[f'I{i}'].value
                currency = sheet[f'J{i}'].value

                if TextileManufact.objects.get_or_create(name=supplier):
                    print(i, supplier, 'существует')

                i += 1

            while y < max_count:
                supplier = sheet[f'A{y}'].value
                collection = sheet[f'B{y}'].value
                designation = sheet[f'D{y}'].value
                color = sheet[f'F{y}'].value
                model = sheet[f'G{y}'].value
                height = sheet[f'H{y}'].value
                price = sheet[f'I{y}'].value
                currency = sheet[f'J{y}'].value

                base_supplier = TextileManufact.objects.get(name=supplier)
                if TextileCollection.objects.get_or_create(manufacturer=base_supplier, name=collection):
                    print(y, collection)
                y += 1

            while z < max_count:
                supplier = sheet[f'A{z}'].value
                collection = sheet[f'B{z}'].value
                designation = sheet[f'D{z}'].value
                color = sheet[f'F{z}'].value
                model = sheet[f'G{z}'].value
                height = sheet[f'H{z}'].value
                price = sheet[f'I{z}'].value
                currency = sheet[f'J{z}'].value

                base_supplier = TextileManufact.objects.get(name=supplier)
                base_collections = TextileCollection.objects.get(manufacturer=base_supplier, name=collection)

                if Textile.objects.get_or_create(manufacturer=base_supplier, collection=base_collections,
                                                 model=transliterate(str(model)), color=color,
                                                 height=height, price_opt=price, designation=designation, currency=currency):
                    instance = Textile.objects.get(manufacturer=base_supplier, collection=base_collections,
                                                   model=transliterate(str(model)), color=color,
                                                   height=height, price_opt=price)
                    instance.article = transliterate(str(collection)[0]).upper() + transliterate(str(model)[0]).upper() + str(instance.pk)
                    instance.save(update_fields=['article'])
                    print(z, instance.article, model, color, height, price)
                z += 1







        # obj = Textile.objects.filter(date_created__gte='2020-10-29')
        #
        #
        # # obj = Textile.objects.filter(manufacturer__name='Дом Каро')
        # for ob in obj:
        #     print(ob.id, ob.article)
        #     ob.delete()
